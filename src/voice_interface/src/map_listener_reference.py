#!/usr/bin/env python  
#coding=utf-8

""" 

- Version 3.0 2015/10/15

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

地图是正的
给机器人提供一个功能性质的api

this code monitor the robot position in map && cell data
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

from openpyxl import load_workbook
from openpyxl import  Workbook
from openpyxl.writer.excel  import  ExcelWriter
from openpyxl.cell import get_column_letter

import getpass,PyKDL,collections,numpy,math
from geometry_msgs.msg import Pose

def decompression(data):
 decompress=eval(data)
 return decompress

def quat_to_angle(quat):
 rot = PyKDL.Rotation.Quaternion(quat[0], quat[1], quat[2], quat[3])
 return rot.GetRPY()[2]

def angle_to_quat(angle):#in rad
 rot = PyKDL.Rotation.RotZ(angle)
 return rot.GetQuaternion()

#real-world (0,0) in matrixs frame即cell坐标
def map_center_cell(map_origin,height,resolution):
 centremap_cell=[0-int(round(map_origin.point.x/resolution)),height+int(round(map_origin.point.y/resolution))]
 return centremap_cell

#返回机器人的cell坐标点格式为[x格数，y格数]
def robot_pose_cell(map_origin,odom_pose,height,resolution):
 centremap_cell=map_center_cell(map_origin,height,resolution)
 robotpose_pose_cell=[centremap_cell[0]+int(round(odom_pose.position.x/resolution)),centremap_cell[1]+int(round(odom_pose.position.y/resolution))]
 return robotpose_pose_cell

#the real world position of a specific point in metrix格式pose（）
def metrix_RealPose(map_origin,goal_cell,goal_ort,height,resolution):
 Goal_Pose=Pose()
 centremap_cell=map_center_cell(map_origin,height,resolution)
 quaternion=angle_to_quat(goal_ort)
 Goal_Pose.position.x=(goal_cell[0]-centremap_cell[0])*resolution#width
 Goal_Pose.position.y=(goal_cell[1]-centremap_cell[1])*resolution#height
 Goal_Pose.orientation.x=quaternion[0]
 Goal_Pose.orientation.y=quaternion[1]
 Goal_Pose.orientation.z=quaternion[2]
 Goal_Pose.orientation.w=quaternion[3]
 return Goal_Pose

#figger out reasonabel goal for robot moving
def robot_goal_compensated(map_origin,goal_cell,goal_ort,height,resolution,data):
 trigger=[True,True]
 checker=0
# 判断x
 #左边
 if trigger[0]:
  for i in range(5):
   if data[goal_cell[1]][goal_cell[0]+i+1]!=0:
    break
   else:
    checker+=1
    if checker==5:
     x=goal_cell[0]+i+1
     trigger[0]=False
  checker=0
 #右边
 if trigger[0]:
  for i in range(5):
   if data[goal_cell[1]][goal_cell[0]-(i+1)]!=0:
    break
   else:
    checker+=1
    if checker==5:
     x=goal_cell[0]-(i+1)
     trigger[0]=False
  checker=0
#判断y
 #左边
 if trigger[1]:
  for j in range(5):
   if data[goal_cell[1]+j+1][x]!=0:
    break
   else:
    checker+=1
    if checker==5:
     y=goal_cell[0]+i+1
     trigger[1]=False
  checker=0
 #右边
 if trigger[1]:
  for j in range(5):
   if data[goal_cell[1]-(j+1)][x]!=0:
    break
   else:
    checker+=1
    if checker==5:
     y=goal_cell[0]-(j+1)
     trigger[1]=False
  checker=0

 robotpose_compen=Pose()
 robotpose_compen=metrix_RealPose(map_origin,[x,y],goal_ort,height,resolution)
 return robotpose_compen



#返回有效区域的坐标集
def effective_point(data):
 map_matrix=map_matrix_ranger(data)
 clear_area,block_area=[],[]
 width=data.info.width
 height=data.info.height
 resolution=data.info.resolution
 map_origin=data.info.origin

 centremap_cell=map_center_cell(map_origin,height,resolution)
 for y in height:
  for x in width:
   if map_matrix[i][j]==0:
    clear_x=(x-centremap_cell[0])*resolution
    clear_y=(y-centremap_cell[1])*resolution
    clear_area.append([clear_x,clear_y])
   elif map_matrix[i][j]==100:
    block_x=(x-centremap_cell[0])*resolution
    block_y=(y-centremap_cell[1])*resolution
    block_area.append([block_x,block_y])
   else:
    pass
 return [clear_area,block_area]


#返回地图矩阵
def map_matrix_ranger(data):
 height=data.info.height
 width=data.info.width
 map_matrix=collections.deque()
 for i in range(height):
  map_width=[]
  for j in range(width):
   map_width.append(data.data[j+width*i])
  map_matrix.appendleft(map_width)
 return map_matrix


#边界历遍程序,检测地图是否完整。
"""checker[[],[]], checker[0]是用来检测所有的y坐标轴是否闭合，checker[1]是用来检测x坐标轴是否闭合。11 表示闭合，00表示敞开，10表示左边单边闭合，01表示右边单边闭合"""
def edge_explorer(data):
 map_matrix=map_matrix_ranger(data)
 height=data.info.height
 checker=[[],[]]
 #测所有x轴
 for y in range(height):
  if 100 in map_matrix[y] and 0 in map_matrix[y]:
   #输出每行 0 和 100 的位置
   block_list=[block_cell for block_cell,block in enumerate(map_matrix[y]) if block==100]
   clear_list=[clear_cell for clear_cell,clear in enumerate(map_matrix[y]) if clear==0]
   #输出该行的状态
   if min(block_list)<min(clear_list):
    if max(clear_list)<max(block_list):
     checker[0].append(11)
   if min(clear_list)<min(block_list):
    if max(block_list)<max(clear_list):
     checker[0].append(00)
   if min(block_list)<min(clear_list):
    if max(block_list)<max(clear_list):
     checker[0].append(10)
   if min(clear_list)<min(block_list):
    if max(clear_list)<max(block_list):
     checker[0].append(01)
  if 100 in map_matrix[y] and 0 not in map_matrix[y]:
   checker[0].append(11)
  if 100 not in map_matrix[y] and 0 in map_matrix[y]:
   checker[0].append(00)
  if 100 not in map_matrix[y] and 0 not in map_matrix[y]:
   checker[0].append(-1)
 #测所有y轴
 map_matrix=numpy.matrix(map_matrix)
 for x in range(width):
  if 100 in map_matrix[:,x] and 0 in map_matrix[:,x]:
   block_list=[block_cell for block_cell,block in enumerate(map_matrix[:,x]) if block==100]
   clear_list=[clear_cell for clear_cell,clear in enumerate(map_matrix[:,x]) if clear==0]
   if min(block_list)<min(clear_list):
    if max(clear_list)<max(block_list):
     checker[1].append(11)
   if min(clear_list)<min(block_list):
    if max(block_list)<max(clear_list):
     checker[1].append(00)
   if min(block_list)<min(clear_list):
    if max(block_list)<max(clear_list):
     checker[1].append(10)
   if min(clear_list)<min(block_list):
    if max(clear_list)<max(block_list):
     checker[1].append(01) 
  if 100 in map_matrix[y] and 0 not in map_matrix[y]:
   checker[0].append(11)
  if 100 not in map_matrix[y] and 0 in map_matrix[y]:
   checker[0].append(00)
  if 100 not in map_matrix[y] and 0 not in map_matrix[y]:
   checker[0].append(-1)

 return checker

#输出一系列点中距离goal最近的一个
def nearest(goal,list_num,diff=0):
 for i in list_num:
  if diff>abs(i-goal):
   diff=abs(i-goal)
   nearest_num=i
 return nearest_num

#全局行动方位决策
def global_action_director(data,odom_pose):
 map_status=edge_explorer(data)

 height=data.info.height
 resolution=data.info.resolution
 map_origin=data.info.origin

 [roobot_ort_x,robot_ort_y]=robot_pose_cell(data.map_origin,odom_pose,height,resolution)

 x_closed=[closed for closed,status in enumerate(map_status[0]) if status==11]
 x_opened=[opened for opened,status in enumerate(map_status[0]) if status==00]
 x_ltopen=[ltopen for ltopen,status in enumerate(map_status[0]) if status==01]
 x_rtopen=[rtopen for rtopen,status in enumerate(map_status[0]) if status==10]
 x_undect=[undect for undect,status in enumerate(map_status[0]) if status==-1]

 y_closed=[closed for closed,status in enumerate(map_status[1]) if status==11]
 y_opened=[opened for opened,status in enumerate(map_status[1]) if status==00]
 y_upopen=[ltopen for ltopen,status in enumerate(map_status[1]) if status==01]
 y_dnopen=[rtopen for rtopen,status in enumerate(map_status[1]) if status==10]
 y_undect=[undect for undect,status in enumerate(map_status[0]) if status==-1]
 
 #确认轴状态完整
 if len(x_closed)+len(x_opened)+len(x_ltopen)+len(x_rtopen)+len(x_undect)==height and len(y_closed)+len(y_opened)+len(y_ltopen)+len(y_rtopen)+len(y_undect)==width:
  x_completion=len(x_closed)/(len(x_closed)+len(x_opened)+len(x_ltopen)+len(x_rtopen))
  y_completion=len(y_closed)/(len(y_closed)+len(y_opened)+len(y_ltopen)+len(y_rtopen))

 #checking 完成度：
  if x_completion<=0.9 or y_completion<=0.9:
  #x轴探索优化
   x_area=x_opened+x_ltopen+x_rtopen
   x_upper_area=[v for i, v in enumerate(x_area) if v<height-roobot_ort_y]
   x_lower_area=[v for i, v in enumerate(x_area) if v>height-roobot_ort_y]
   x_detector_area={'upper':x_upper_area,'lower':x_lower_area}
   if abs(len(x_upper_area)-len(x_lower_area))>=10:
    if len(x_upper_area)>=len(x_lower_area):
     detector_area_x='upper'
    else:
     detector_area_x='lower'
    goal_y=nearest(roobot_ort_y,x_detector_area[detector_area_x])
   else:
    goal_upper_y=nearest(roobot_ort_y,x_detector_area['upper'])
    goal_lower_y=nearest(roobot_ort_y,x_detector_area['lower'])
    if abs(goal_upper_y-roobot_ort_y)<abs(goal_lower_y-roobot_ort_y):
     goal_y=goal_upper_y
     detector_area_x='upper'
    elif abs(goal_upper_y-roobot_ort_y)>abs(goal_lower_y-roobot_ort_y):
     goal_y=goal_lower_y
     detector_area_x='lower'
    else:
     print '上下区域距离相等，或者有未考虑的情况'
     goal_y=goal_upper_y
     detector_area_x='upper'
  #y轴探索优化
   y_area=y_opened+y_ltopen+y_rtopen
   y_left_area=[v for i, v in enumerate(y_area) if v<roobot_ort_x]
   y_right_area=[v for i, v in enumerate(y_area) if v>roobot_ort_x]
   y_detector_area={'left':x_left_area,'right':x_right_area}
   if abs(len(y_left_area)-len(y_right_area))>=10:
    if len(y_left_area)>=len(y_right_area):
     detector_area_y='left'
    else:
     detector_area_y='right'
    goal_x=nearest(roobot_ort_x,y_detector_area[detector_area_y])
   else:
    goal_left_x=nearest(roobot_ort_x,y_detector_area['left'])
    goal_right_x=nearest(roobot_ort_x,y_detector_area['right'])
    if abs(goal_left_x-roobot_ort_x)<abs(goal_right_x-roobot_ort_x):
     goal_x=goal_left_x
     detector_area_y='left'
    elif abs(goal_left_x-roobot_ort_x)>abs(goal_right_x-roobot_ort_x):
     goal_x=goal_right_x
     detector_area_y='right'
    else:
     print '左右区域距离相等，或者有未考虑的情况'
     goal_x=goal_left_x
     detector_area_x='upper'
   return [[detector_area_x,detector_area_y],[goal_x,goal_y]]
  else:
   return [['complete','complete'],[roobot_ort_x,roobot_ort_y]]
 else:
  return False

#存储 for testing convenient
def store_xlsx(data,high,width):
 accout=getpass.getuser()
 excel=Workbook()
 excel_writer=ExcelWriter(workbook = excel)
 excel_outpath= r'/home/%s/mapdata/test1.xlsx'%accout
 sheet = excel.create_sheet() 
 sheet=excel.worksheets[0]
 sheet.title='testmap'

 for row in range (1,(high+1)):
  for col in range (1,(width+1)):
   column = get_column_letter(col)
   sheet.cell('%s%s'%(column,row)).value='%s'%data[row-1][col-1]
   
 for col in range(1,width+1):
  column = get_column_letter(col)
  sheet.column_dimensions['%s'%column].width = 2.6
 for row in range(1,high+1):
  sheet.row_dimensions[row].height =14.15

 sheet = excel.create_sheet()
 sheet.title='parameters'
 sheet.cell('A1').value='parameters'
 sheet.cell('A2').value='This represents a 2-D grid map, in which each cell represents the probability of occupancy.'

 sheet.cell('A7').value='Map width [cells]'
 sheet.cell('A8').value='%s'%width
 sheet.cell('C7').value='Map height [cells]'
 sheet.cell('C8').value='%s'%high

 excel.save(filename=excel_outpath)
 print 'saving process done'
 return 'ok'
#存txt
def store_txt(data):
 store=open('./cellmap.txt','w')	
 for i in range(n):
  store.writeline('%d '%data.data[i])
 store.write('\n')
 print 'written int cellmap.txt file'
 store.close()
 return 'ok'
 
# 产生一个朝向当前目标点的角度
def angle_generater(sub_point,pre_point):
 if pre_point.point.x<=sub_point.point.x:
  angle= math.atan((sub_point.point.y-pre_point.point.y)/(sub_point.point.x-pre_point.point.x))
 if pre_point.point.x>sub_point.point.x:
  angle= math.pi+math.atan((sub_point.point.y-pre_point.point.y)/(sub_point.point.x-pre_point.point.x))
 return angle
