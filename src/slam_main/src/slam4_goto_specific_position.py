#!/usr/bin/env python
""" 
- Version 1.0 2015/10/15
 
Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

go to specific position in map and back(worked,done)

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

rostopic pub /move_base_simple/goal geometry_msgs/PoseStamped '{ header: { frame_id: "map" }, pose: { position: { x: 1.0, y: 0, z: 0 }, orientation: { x: 0, y: 0, z: 0, w: 1 } } }'

need to os.system('export TURTLEBOT_MAP_FILE=~/maps/map_%d.yaml'%self.count)#export map

this file work with helping of following files:

utils/execl_coordinate_without_positive_coordinate.py
utils/coordinate_execl_without_positive_coordinate.py
robot_status/map_lisenter_without_positive_coordinate.py
gmapping
"""
import rospy, actionlib, os, getpass,math
from actionlib_msgs.msg import *
from geometry_msgs.msg import Pose
from nav_msgs.msg import Odometry
from move_base_msgs.msg import MoveBaseAction, MoveBaseGoal
from std_msgs.msg import String

from openpyxl import load_workbook
from openpyxl import  Workbook
from openpyxl.writer.excel  import  ExcelWriter
#from openpyxl.cell import get_column_letter

class goto_specified_position():

 def goals(self,goal):
  if goal.data!='start':
   print 'goal is :',goal.data
   self.action=actionlib.SimpleActionClient('move_base',MoveBaseAction)
   self.action.wait_for_server(rospy.Duration(60))

   self.goal=MoveBaseGoal()
   self.goal.target_pose.pose.position.x=float(goal.data.split(',')[0])
   self.goal.target_pose.pose.position.y=float(goal.data.split(',')[1])
   self.goal.target_pose.pose.orientation.w=1
   self.goal.target_pose.header.frame_id="map"
   self.goal.target_pose.header.stamp=rospy.Time.now()

   print self.goal
   self.action.send_goal(self.goal)
   finish_time=self.action.wait_for_result()#rospy.Duration(240))
   state=self.action.get_state()
   robot_inital=Pose()
   robot_inital=self.odoms()
   goal_states={
	'%s'%GoalStatus.PENDING:'PENDING',
	'%s'%GoalStatus.ACTIVE:'ACTIVE',
	'%s'%GoalStatus.SUCCEEDED:'SUCCEEDED',
	'%s'%GoalStatus.PREEMPTED:'PREEMPTED',
	'%s'%GoalStatus.ABORTED:'ABORTED',
	'%s'%GoalStatus.REJECTED:'REJECTED',
	'%s'%GoalStatus.PREEMPTING:'PREEMPTING',
	'%s'%GoalStatus.RECALLING:'RECALLING',
	'%s'%GoalStatus.RECALLED:'RECALLED',
	'%s'%GoalStatus.LOST:'LOST'}
   print goal_states

   if not finish_time:
    self.action.cancel_goal()
    print state,'out of time'
   if state==GoalStatus.SUCCEEDED:
    print state,'goal is currently achieved'
    print 'going back'


    self.odoms_store(robot_inital)

    print 'home coordinate:\n',robot_inital
    self.goal.target_pose.pose.position=robot_inital.position
    self.goal.target_pose.pose.orientation=robot_inital.orientation
    self.action.send_goal(self.goal)
    back_time=self.action.wait_for_result()#rospy.Duration(240))
    state=self.action.get_state()
    if not back_time :
     self.action.cancel_goal()
     print state,'out of time'
    if state==GoalStatus.SUCCEEDED:
     print state,'home is currently achieved'
     self.odoms_store(robot_inital)
    else:
     print 'cannot be achieved the home due to:', state,str(goal_states['%s'%state])
     self.odoms_store(robot_inital)

   else:
    print 'cannot be achieved the goal due to:', state, str(goal_states['%s'%state])
    self.odoms_store(robot_inital)
    print 'going back'

    robot_inital=self.odoms()

    self.odoms_store(robot_inital)
    print 'home coordinate:\n',robot_inital
    self.goal.target_pose.pose.position=robot_inital.position
    self.goal.target_pose.pose.orientation=robot_inital.orientation
    self.action.send_goal(self.goal)
    back_time=self.action.wait_for_result()#rospy.Duration(240))
    state=self.action.get_state()
    if not back_time :
     self.action.cancel_goal()
     print state,'out of time'
    if state==GoalStatus.SUCCEEDED:
     print state,'home is currently achieved'
     self.odoms_store(robot_inital)
    else:
     print 'cannot be achieved the home due to:', state,str(goal_states['%s'%state])
     self.odoms_store(robot_inital)

    self.travel_distance= math.sqrt(pow((self.current_odom.pose.pose.position.x-robot_inital.position.x),2)+pow((self.current_odom.pose.pose.position.y-robot_inital.position.y),2))
    pub = rospy.Publisher('travel_distance', String, queue_size=5)
    pub.publish('%s'%str(self.travel_distance))


 def odoms(self):
  robot_inital=Pose()
  read_path=r'/home/%s/mapdata/cellmap.xlsx'%self.accout
  wb = load_workbook(filename = read_path )
  sheetnames=wb.get_sheet_names()
  ws=wb.get_sheet_by_name('parameters')
  robot_inital.position.x=float(ws.cell('A16').value)
  robot_inital.position.y=float(ws.cell('B16').value)
  robot_inital.position.z=float(ws.cell('C16').value)
  robot_inital.orientation.x=float(ws.cell('A18').value)
  robot_inital.orientation.y=float(ws.cell('B18').value)
  robot_inital.orientation.z=float(ws.cell('C18').value)
  robot_inital.orientation.w=float(ws.cell('D18').value)
  print 'reading from data base:\n',robot_inital
  return robot_inital

 def odoms_store(self,frist_pose):
  self.robot_odom=Pose()
  excel=Workbook()
  excel_writer=ExcelWriter(workbook = excel)
  store_path=r'/home/%s/mapdata/odoms.xlsx'%self.accout
  sheet = excel.create_sheet() 
  sheet=excel.worksheets[0]
  sheet.title='odoms'
  sheet.cell('A1').value='odmos'
  sheet.cell('A2').value='initial pose (x,y,z), orientation (x,y,z,w):'
  sheet.cell('A3').value=0.4#frist_pose.position.x
  sheet.cell('B3').value=-0.05#frist_pose.position.y
  sheet.cell('C3').value=0#frist_pose.position.z
  sheet.cell('E3').value=frist_pose.orientation.x
  sheet.cell('F3').value=frist_pose.orientation.y
  sheet.cell('G3').value=frist_pose.orientation.z
  sheet.cell('H3').value=frist_pose.orientation.w

  if self.current_pose.pose.pose.position != self.current_odom.pose.pose.position or self.current_pose.pose.pose.orientation!=self.current_odom.pose.pose.orientation:
   sheet.cell('A%s'%(1+5*self.count)).value='current pose (x,y,z), orientation (x,y,z,w):'
   sheet.cell('A%s'%(2+5*self.count)).value=self.current_odom.pose.pose.position.x
   sheet.cell('B%s'%(2+5*self.count)).value=self.current_odom.pose.pose.position.y
   sheet.cell('C%s'%(2+5*self.count)).value=self.current_odom.pose.pose.position.z

   sheet.cell('E%s'%(3+5*self.count)).value=self.current_odom.pose.pose.orientation.x
   sheet.cell('F%s'%(3+5*self.count)).value=self.current_odom.pose.pose.orientation.y
   sheet.cell('G%s'%(3+5*self.count)).value=self.current_odom.pose.pose.orientation.z
   sheet.cell('H%s'%(3+5*self.count)).value=self.current_odom.pose.pose.orientation.w

   self.current_pose.pose.pose.position=self.current_odom.pose.pose.position
   self.current_pose.pose.pose.orientation=self.current_odom.pose.pose.orientation
   #self.count+=1
  else:
   pass

  excel.save(filename=store_path)

 def tfcallback(self,data):
  self.current_pose=Odometry()
  self.current_odom=Odometry()
  self.current_odom=data

 def __init__(self):
  self.count=1
  self.travel_distance=0.0
  self.accout=getpass.getuser()
  self.robot_odom=Pose()
  print "start slamming"
  rospy.init_node('goto_specified_position', anonymous=False)
  rospy.wait_for_message('excel_to_coordinate', String)
  rospy.Subscriber('excel_to_coordinate', String, self.goals)
  rospy.Subscriber("/odom", Odometry, self.tfcallback)
  rospy.spin()
  
if __name__=='__main__':
 try:
  print "initialization system"
  goto_specified_position()
  print "process done and quit"
 except rospy.ROSInterruptException:
  rospy.loginfo("goto specified position node terminated.")
