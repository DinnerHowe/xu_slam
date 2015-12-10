#!/usr/bin/env python  
""" 

- Version 3.0 2015/10/12

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

this code monitor the robot position in map && cell data
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

import rospy , os
import xlwt, xlrd
import csv
from openpyxl import load_workbook
from openpyxl import  Workbook
from openpyxl.writer.excel  import  ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill

from nav_msgs.msg import OccupancyGrid
from robot_status.msg import mapdata
#from robot_status.msg import robot_odom
from geometry_msgs.msg import PoseWithCovarianceStamped
import getpass

class listener():
 def store_celloccupancy(self,data):
  accout=getpass.getuser()
  store=open('/home/%s/mapdata/cellmap.txt'%accout,'w')	
  n=len(data.data)
  data.data+=('last',)
  print n,len(data.data),self.num #data.data,'done'
  for i in range(n):
   store.write('%d '%data.data[i])
   if i%288==0 and i!=0:
    store.write('\n')
  print 'written int cellmap.txt file'
  store.close()

 def store_csv(self,data):
  accout=getpass.getuser()
  n=len(data.data)
  data.data+=('last',)
  print n,len(data.data),self.num 
  tempdict={}
  templist=[]
  for row in range(n/288):
   for col in range(288):
    templist.append('%d'%data.data[row*288+col])
   tempdict['%s'%row]=templist
  tempdict['done']='done'
  with open('/home/%s/mapdata/cellmap.csv'%accout,'wb') as csvfile:
   csv_writer=csv.writer(csvfile, dialect='excel')
   for i in tempdict:
    csv_writer.writerow(tempdict[i])
  print 'written in to cellmap.csv file'

 def store_xls(self,data):
  accout=getpass.getuser()
  excel=xlwt.Workbook()
  n=len(data.data)
  data.data+=('last',)
  print n,len(data.data),self.num 
  i=0
  marker=0
  sheet=excel.add_sheet('cellmap')
  for col in range(0,255):
   sheet.write(0,col,marker)
   marker+=1
  for row in range (1,int(n/255)+1):
   for col in range (0,255):
    sheet.write(row,col,'%s'%data.data[i])
    i+=1
  excel.save('/home/%s/mapdata/cellmap.xls'%accout)
  book=xlrd.open_workbook('/home/%s/mapdata/cellmap.xls'%accout)
  sh = book.sheet_by_index(0)
  print 'written in to cellmap.xls file'
  print "Worksheet name:", book.sheet_names()
  print 'sheet name:', sh.name, 'rows:', sh.nrows, 'cols:',sh.ncols

 def checker(self,trigger,i,data,n):
  nthread=''
  if trigger=='check':
   nthread='done'
   print 'n-i-1',((len(data)-1)-i),data[(len(data)-1)-i],'\n','n-i-2',((len(data)-2)-i),data[(len(data)-2)-i]
  if trigger=='':
   print 'n-i',data[(n-i)],'n:', n,'i:',i
   nthread='check'
  return nthread

 def colormark(self,column,height,i,data,sheet):
  rospy.loginfo('coloring map, it might take for a while')
  sheet.cell('%s%s'%(column,height)).style.fill.fill_type =Fill.FILL_SOLID
  if data[i]==100:
   sheet.cell('%s%s'%(column,height)).style.fill.start_color.index =Color.RED
  if data[i]==0:
   sheet.cell('%s%s'%(column,height)).style.fill.start_color.index =Color.BLUE

 def store_xlsx(self,data):
  accout=getpass.getuser()
  excel=Workbook()#optimized_write = True)
  excel_writer=ExcelWriter(workbook = excel)
  excel_outpath= r'/home/%s/mapdata/cellmap.xlsx'%accout
  sheet = excel.create_sheet() 
  sheet=excel.worksheets[0]
  sheet.title='cellmap_%d'%self.num

  wb = load_workbook(filename = excel_outpath )
  sheetnames=wb.get_sheet_names()
  ws=wb.get_sheet_by_name('cellmap_%d'%self.num)
  
  n=len(data.data)
  data.data+=('last',)
  print 'data length:',n,'data,length with ending mark:',len(data.data),self.num 
  sheet.cell('A1').value='mapdata'
  marker=0
  sheet.column_dimensions['A'].width = 1.89
  sheet.row_dimensions[1].height =14.15
  for col in range(2,data.info.width+2):
   column = get_column_letter(col)
   sheet.cell('%s%s'%(column,1)).value='%d'%marker
   sheet.column_dimensions['%s'%column].width = 1.89
   marker+=1
  marker=0
  for row in range(2,data.info.height+2):
   sheet.cell('%s%s'%('A',data.info.height+3-row)).value='%d'%marker
   sheet.row_dimensions[(data.info.height+3-row)].height =14.15
   marker+=1
  i=0
  thread=''
  for row in range (2,(data.info.height+2)):
   for col in range (2,(data.info.width+2)):
    column = get_column_letter(col)
    #thread=self.checker(thread,i,data.data,n)
    #thread=self.checker(thread,i,data.data,n)
    height=(data.info.height+2)-row+1
    sheet.cell('%s%s'%(column,height)).value='%s'%data.data[i]
    #self.colormark(column,height,i,data.data,sheet)
    i+=1
  sheet = excel.create_sheet()
  sheet.title='parameters'
  sheet.cell('A1').value='parameters'
  sheet.cell('A2').value='This represents a 2-D grid map, in which each cell represents the probability of occupancy.'
  sheet.cell('A4').value='map_load_time'
  sheet.cell('A5').value='%s'%data.info.map_load_time
  sheet.cell('C4').value='The map resolution [m/cell]'
  sheet.cell('C5').value='%s'%data.info.resolution
  sheet.cell('A7').value='Map width [cells]'
  sheet.cell('A8').value='%s'%data.info.width
  sheet.cell('C7').value='Map height [cells]'
  sheet.cell('C8').value='%s'%data.info.height
  sheet.cell('A10').value='map centre MAP(0,0)'
  sheet.cell('A10').value='map centre MAP(0,0), cell(0,0)=B%s'%ws.get_highest_row() 
  sheet.cell('A11').value='position(x,y,z):'
  sheet.cell('C11').value='%s'%data.info.origin.position.x
  sheet.cell('D11').value='%s'%data.info.origin.position.y
  sheet.cell('E11').value='%s'%data.info.origin.position.z
  sheet.cell('A12').value='orientation(x,y,z,w):'
  sheet.cell('C12').value='%s'%data.info.origin.orientation.x
  sheet.cell('D12').value='%s'%data.info.origin.orientation.y
  sheet.cell('E12').value='%s'%data.info.origin.orientation.z
  sheet.cell('F12').value='%s'%data.info.origin.orientation.w
  sheet.cell('A14').value='robot odom'
  sheet.cell('A15').value='robot position: (x ,y ,z )'
  sheet.cell('A16').value='%s'%self.robot_odom.pose.pose.position.x
  sheet.cell('B16').value='%s'%self.robot_odom.pose.pose.position.y
  sheet.cell('C16').value='%s'%self.robot_odom.pose.pose.position.z

  sheet.cell('A17').value='robot rotation: (x ,y ,z ,w )'
  sheet.cell('A18').value='%s'%self.robot_odom.pose.pose.orientation.x
  sheet.cell('B18').value='%s'%self.robot_odom.pose.pose.orientation.y
  sheet.cell('C18').value='%s'%self.robot_odom.pose.pose.orientation.z
  sheet.cell('D18').value='%s'%self.robot_odom.pose.pose.orientation.w

  sheet.cell('A19').value='current time:'
  sheet.cell('A20').value='%s'%str(self.robot_odom.header.stamp)
  sheet.cell('A21').value='robot frame_id:'
  sheet.cell('A22').value='%s'%self.robot_odom.header.frame_id

  excel.save(filename=excel_outpath)
  print 'written in to excel file cellmap.xlsx'

  print "Worksheet range(s):", wb.get_named_ranges()  
  print "Worksheet name(s):", wb.get_sheet_names()  

  print "Worksheet title:", ws.title
  print 'cols:', ws.get_highest_column(), 'rows:', ws.get_highest_row() 
  print 'width:', data.info.width,'0~%d'%(data.info.width-1), 'height:', data.info.height,'0~%d'%(data.info.height-1)

 def callback(self,data):
  Ocdata=OccupancyGrid()
  currentdata=mapdata()
  print 'current map info: \n', data.info
  pub = rospy.Publisher('robot_map', mapdata, queue_size=10)
  Ocdata=data
  #self.store_celloccupancy(Ocdata) #uncomman this to save into *.txt
  #self.store_csv(Ocdata)  #uncomman this to save into *.csv
  #self.store_xls(Ocdata)  #uncomman this to save into *.xls
  self.store_xlsx(Ocdata)
  currentdata.grid=Ocdata
  currentdata.mapnum=self.num
  currentdata.odom.position=self.robot_odom.pose.pose.position
  currentdata.odom.rotation=self.robot_odom.pose.pose.orientation
  currentdata.odom.header=self.robot_odom.header
  os.system('rosrun map_server map_saver -f /tmp/my_map_0')
  pub.publish(currentdata)
  self.num+=1
  print 'process done'
  '''signal=raw_input('continue?:\n')
  if signal=='' or signal=='y' or signal=='yes':
   pass
  else:
   os.system('rosnode kill map_lisenter')'''
  os.system('rosnode kill map_lisenter')
  
 def tfcallback(self,data):
  self.robot_odom=PoseWithCovarianceStamped()
  self.robot_odom=data
  #rospy.loginfo(self.robot_odom)

 def __init__(self):
  self.num=0
  print "start listening to map"
  rospy.init_node('map_lisenter', anonymous=False)
  #rospy.Subscriber("robot_odom", robot_odom, self.tfcallback)
  rospy.Subscriber("robot_pose_ekf/odom_combined", PoseWithCovarianceStamped, self.tfcallback)
  rospy.Subscriber("map", OccupancyGrid, self.callback)
  rospy.spin()

if __name__ == '__main__':
 listener()

 
