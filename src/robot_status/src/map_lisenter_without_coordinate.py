#!/usr/bin/env python  
#coding=utf-8

""" 

- Version 3.0 2015/10/15

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

地图是反的，但是其他都没问题

this code monitor the robot position in map && cell data
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

import rospy , os
import xlwt, xlrd
import csv
from openpyxl import load_workbook
from openpyxl import  Workbook
from openpyxl.writer.excel  import  ExcelWriter
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill

from nav_msgs.msg import OccupancyGrid
from robot_status.msg import mapdata
from geometry_msgs.msg import PoseWithCovarianceStamped
import getpass

class listener():

 def checker(self,trigger,i,data,n):
  nthread=''
  if trigger=='check':
   nthread='done'
   print 'n-i-1',((len(data)-1)-i),data[(len(data)-1)-i],'\n','n-i-2',((len(data)-2)-i),data[(len(data)-2)-i]
  if trigger=='':
   print 'n-i',data[(n-i)],'n:', n,'i:',i
   nthread='check'
  return nthread

 def colormark(self,column,height,i,data,sheet):
  rospy.loginfo('coloring map, it might take for a while')
  sheet.cell('%s%s'%(column,height)).style.fill.fill_type =Fill.FILL_SOLID
  if data[i]==100:
   sheet.cell('%s%s'%(column,height)).style.fill.start_color.index =Color.RED
  if data[i]==0:
   sheet.cell('%s%s'%(column,height)).style.fill.start_color.index =Color.BLUE

 def store_xlsx(self,data):
  accout=getpass.getuser()
  excel=Workbook()
  excel_writer=ExcelWriter(workbook = excel)
  excel_outpath= r'/home/%s/mapdata/cellmap.xlsx'%accout
  sheet = excel.create_sheet() 
  sheet=excel.worksheets[0]
  sheet.title='cellmap'
  n=len(data.data)
  print 'data length:',n,'data length:',len(data.data) 
  i=0
  thread=''
  for row in range (1,(data.info.height+1)):
   for col in range (1,(data.info.width+1)):
    column = get_column_letter(col)
    #thread=self.checker(thread,i,data.data,n)
    #thread=self.checker(thread,i,data.data,n)
    #height=(data.info.height+2)-row
    height=row
    sheet.cell('%s%s'%(column,height)).value='%s'%data.data[i]
    #self.colormark(column,height,i,data.data,sheet)
    i+=1
  for col in range(1,data.info.width+1):
   column = get_column_letter(col)
   #sheet.column_dimensions['%s'%column].width = 1.89
   sheet.column_dimensions['%s'%column].width = 2.6
  for row in range(1,data.info.height+1):
   sheet.row_dimensions[data.info.height].height =14.15

  sheet = excel.create_sheet()
  sheet.title='parameters'
  sheet.cell('A1').value='parameters'
  sheet.cell('A2').value='This represents a 2-D grid map, in which each cell represents the probability of occupancy.'
  sheet.cell('A4').value='map_load_time'
  sheet.cell('A5').value='%s'%data.info.map_load_time
  sheet.cell('C4').value='The map resolution [m/cell]'
  sheet.cell('C5').value='%s'%data.info.resolution
  sheet.cell('A7').value='Map width [cells]'
  sheet.cell('A8').value='%s'%data.info.width
  sheet.cell('C7').value='Map height [cells]'
  sheet.cell('C8').value='%s'%data.info.height
  sheet.cell('A10').value='map centre MAP(0,0)' 
  sheet.cell('A11').value='position(x,y,z):'
  sheet.cell('C11').value='%s'%data.info.origin.position.x
  sheet.cell('D11').value='%s'%data.info.origin.position.y
  sheet.cell('E11').value='%s'%data.info.origin.position.z
  sheet.cell('A12').value='orientation(x,y,z,w):'
  sheet.cell('C12').value='%s'%data.info.origin.orientation.x
  sheet.cell('D12').value='%s'%data.info.origin.orientation.y
  sheet.cell('E12').value='%s'%data.info.origin.orientation.z
  sheet.cell('F12').value='%s'%data.info.origin.orientation.w
  sheet.cell('A14').value='robot odom'
  sheet.cell('A15').value='robot position: (x ,y ,z )'
  sheet.cell('A16').value='%s'%self.robot_odom.pose.pose.position.x
  sheet.cell('B16').value='%s'%self.robot_odom.pose.pose.position.y
  sheet.cell('C16').value='%s'%self.robot_odom.pose.pose.position.z

  sheet.cell('A17').value='robot rotation: (x ,y ,z ,w )'
  sheet.cell('A18').value='%s'%self.robot_odom.pose.pose.orientation.x
  sheet.cell('B18').value='%s'%self.robot_odom.pose.pose.orientation.y
  sheet.cell('C18').value='%s'%self.robot_odom.pose.pose.orientation.z
  sheet.cell('D18').value='%s'%self.robot_odom.pose.pose.orientation.w

  sheet.cell('A19').value='current time:'
  sheet.cell('A20').value='%s'%str(self.robot_odom.header.stamp)
  sheet.cell('A21').value='robot frame_id:'
  sheet.cell('A22').value='%s'%self.robot_odom.header.frame_id

  excel.save(filename=excel_outpath)
  print 'written in to excel file cellmap.xlsx'

  wb = load_workbook(filename = excel_outpath )
  sheetnames=wb.get_sheet_names()
  ws=wb.get_sheet_by_name(sheetnames[0]) 
  print "Worksheet range(s):", wb.get_named_ranges()  
  print "Worksheet name(s):", wb.get_sheet_names()  

  print "Worksheet title:", ws.title
  print 'cols:', ws.get_highest_column(), 'rows:', ws.get_highest_row() 
  print 'width:', data.info.width,'0~%d'%(data.info.width-1), 'height:', data.info.height,'0~%d'%(data.info.height-1)

 def callback(self,data):
  Ocdata=OccupancyGrid()
  currentdata=mapdata()
  print 'current map info: \n', data.info
  pub = rospy.Publisher('robot_map', mapdata, queue_size=10)
  Ocdata=data
  self.store_xlsx(Ocdata)
  currentdata.grid=Ocdata
  currentdata.mapnum=self.num
  currentdata.odom.position=self.robot_odom.pose.pose.position
  currentdata.odom.rotation=self.robot_odom.pose.pose.orientation
  currentdata.odom.header=self.robot_odom.header
  os.system('rosrun map_server map_saver -f ~/maps/map')
  pub.publish(currentdata)
  self.num+=1
  print 'process done'
  os.system('rosnode kill map_lisenter')
  
 def tfcallback(self,data):
  self.robot_odom=PoseWithCovarianceStamped()
  self.robot_odom=data
  #rospy.loginfo(self.robot_odom)

 def __init__(self):
  self.num=0
  print "start listening to map"
  rospy.init_node('map_lisenter', anonymous=False)
  #rospy.Subscriber("robot_odom", robot_odom, self.tfcallback)
  rospy.Subscriber("robot_pose_ekf/odom_combined", PoseWithCovarianceStamped, self.tfcallback)
  rospy.Subscriber("map", OccupancyGrid, self.callback)
  rospy.spin()

if __name__ == '__main__':
 listener()

 
