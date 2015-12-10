#!/usr/bin/env python  
#coding=utf-8

""" 

- Version 3.0 2015/10/15

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

地图是正的
给机器人提供一个功能性质的api

this code monitor the robot position in map && cell data
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

from openpyxl import load_workbook
from openpyxl import  Workbook
from openpyxl.writer.excel  import  ExcelWriter
from openpyxl.cell import get_column_letter

import getpass,PyKDL,collections
from geometry_msgs.msg import Pose
from math import *

def decompression(data):
 decompress=eval(data)
 return decompress

def ort_pose(yaw,pose,r):
 x=sin(yaw)*r+pose.x
 y=cos(yaw)*r+pose.y
 return [x,y]

def quat_to_angle(quat):
 rot = PyKDL.Rotation.Quaternion(quat.x, quat.y, quat.z, quat.w)
 return rot.GetRPY()[2]

def angle_to_quat(angle):#in rad
 rot = PyKDL.Rotation.RotZ(angle)
 return rot.GetQuaternion()

#real-world (0,0) in matrixs frame即cell坐标
def map_center_cell(map_origin,height,resolution):
 centremap_cell=[0-int(round(map_origin.position.x/resolution)),height+int(round(map_origin.position.y/resolution))]
 return centremap_cell

#返回机器人的cell坐标点格式为[x格数，y格数]
def robot_pose_cell(map_origin,odom_pose,height,resolution):
 centremap_cell=map_center_cell(map_origin,height,resolution)
 robotpose_pose_cell=[centremap_cell[0]+int(round(odom_pose.position.x/resolution)),centremap_cell[1]+int(round(odom_pose.position.y/resolution))]
 return robotpose_pose_cell

#the real world position of a specific point in metrix格式pose（）
def metrix_RealPose(map_origin,goal_cell,goal_ort,height,resolution):
 Goal_Pose=Pose()
 centremap_cell=map_center_cell(map_origin,height,resolution)
 quaternion=angle_to_quat(goal_ort)
 Goal_Pose.position.x=(goal_cell[0]-centremap_cell[0])*resolution#width
 Goal_Pose.position.y=(goal_cell[1]-centremap_cell[1])*resolution#height
 Goal_Pose.orientation.x=quaternion[0]
 Goal_Pose.orientation.y=quaternion[1]
 Goal_Pose.orientation.z=quaternion[2]
 Goal_Pose.orientation.w=quaternion[3]
 return Goal_Pose

#figger out reasonabel goal for robot moving

def robot_goal_compensated(map_origin,goal_cell,goal_ort,height,resolution,data):
 trigger=[True,True]
 checker=0
# 判断x
 if trigger[0]:
  for i in range(5):
   if data[goal_cell[1]][goal_cell[0]+i+1]!=0:
    break
   else:
    checker+=1
    if checker==5:
     x=goal_cell[0]+i+1
     trigger[0]=False
  checker=0

 if trigger[0]:
  for i in range(5):
   if data[goal_cell[1]][goal_cell[0]-(i+1)]!=0:
    break
   else:
    checker+=1
    if checker==5:
     x=goal_cell[0]-(i+1)
     trigger[0]=False
  checker=0

#判断y
 if trigger[1]:
  for j in range(5):
   if data[goal_cell[1]+j+1][x]!=0:
    break
   else:
    checker+=1
    if checker==5:
     y=goal_cell[0]+i+1
     trigger[1]=False
  checker=0

 if trigger[1]:
  for j in range(5):
   if data[goal_cell[1]-(j+1)][x]!=0:
    break
   else:
    checker+=1
    if checker==5:
     y=goal_cell[0]-(j+1)
     trigger[1]=False
  checker=0

 robotpose_compen=Pose()
 robotpose_compen=metrix_RealPose(map_origin,[x,y],goal_ort,height,resolution)
 return robotpose_compen

#边界历遍程序
def edge_explorer(data):
 map_matrix=map_matrix_ranger(data)
 height=data.info.height
 for y in range(height):
  if 100 in map_matrix[y] or 0 in map_matrix[y]:
   block_list=[block_cell for block_cell,block in enumerate(map_matrix[y]) if block==100]
   clear_list=[clear_cell for clear_cell,clear in enumerate(map_matrix[y]) if clear==0]
   if min(block_list)<min(clear_list):
    if max(block_list)>max(clear_list):
     area=True




#返回有效区域的坐标集data是订阅map的完整数据
def effective_point(data):
 map_matrix=map_matrix_ranger(data)
 clear_area,block_area=[],[]
 width=data.info.width
 height=data.info.height
 resolution=data.info.resolution
 map_origin=data.info.origin

 centremap_cell=map_center_cell(map_origin,height,resolution)
 for y in range(height):
  for x in range(width):
   if map_matrix[y][x]==0:
    clear_x=(x-centremap_cell[0])*resolution
    clear_y=(y-centremap_cell[1])*resolution
    clear_area.append([clear_x,clear_y])
   elif map_matrix[y][x]==100:
    block_x=(x-centremap_cell[0])*resolution
    block_y=(y-centremap_cell[1])*resolution
    block_area.append([block_x,block_y])
   else:
    pass
 return [clear_area,block_area]

#返回有效区域的cell坐标集data是订阅map的完整数据
def effective_point_cell(data):
 map_matrix=map_matrix_ranger(data)
 clear_area,block_area=[],[]
 width=data.info.width
 height=data.info.height
 resolution=data.info.resolution
 map_origin=data.info.origin

 centremap_cell=map_center_cell(map_origin,height,resolution)
 for y in range(height):
  for x in range(width):
   if map_matrix[y][x]==0:
    clear_x=(x-centremap_cell[0])
    clear_y=(y-centremap_cell[1])
    clear_area.append([clear_x,clear_y])
   elif map_matrix[y][x]==100:
    block_x=(x-centremap_cell[0])
    block_y=(y-centremap_cell[1])
    block_area.append([block_x,block_y])
   else:
    pass
 return [clear_area,block_area]

#返回地图矩阵
def map_matrix_ranger(data):
 height=data.info.height
 width=data.info.width
 map_matrix=collections.deque()
 for i in range(height):
  map_width=[]
  for j in range(width):
   map_width.append(data.data[j+width*i])
  map_matrix.appendleft(map_width)
 return map_matrix


#存储 for testing convenient
def store_xlsx(data,high,width):
 accout=getpass.getuser()
 excel=Workbook()
 excel_writer=ExcelWriter(workbook = excel)
 excel_outpath= r'/home/%s/mapdata/test1.xlsx'%accout
 sheet = excel.create_sheet() 
 sheet=excel.worksheets[0]
 sheet.title='testmap'

 for row in range (1,(high+1)):
  for col in range (1,(width+1)):
   column = get_column_letter(col)
   sheet.cell('%s%s'%(column,row)).value='%s'%data[row-1][col-1]
   
 for col in range(1,width+1):
  column = get_column_letter(col)
  sheet.column_dimensions['%s'%column].width = 2.6
 for row in range(1,high+1):
  sheet.row_dimensions[row].height =14.15

 sheet = excel.create_sheet()
 sheet.title='parameters'
 sheet.cell('A1').value='parameters'
 sheet.cell('A2').value='This represents a 2-D grid map, in which each cell represents the probability of occupancy.'

 sheet.cell('A7').value='Map width [cells]'
 sheet.cell('A8').value='%s'%width
 sheet.cell('C7').value='Map height [cells]'
 sheet.cell('C8').value='%s'%high

 excel.save(filename=excel_outpath)
 print 'saving process done'
 return 'ok'

def store_txt(data):
 store=open('./cellmap.txt','w')	
 for i in range(n):
  store.writeline('%d '%data.data[i])
 store.write('\n')
 print 'written int cellmap.txt file'
 store.close()

 
