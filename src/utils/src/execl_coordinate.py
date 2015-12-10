#!/usr/bin/env python
#coding=utf-8
""" 

- Version 2.0 2015/10/13

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

this code transfer execl fram to coordinate fram
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

import rospy
from openpyxl.cell import get_column_letter
from std_msgs.msg import String
import string
from openpyxl import load_workbook
import getpass

class excel_to_coordinate():

 def __init__(self):
  rospy.init_node('excel_to_coordinate',anonymous=True)
  pub = rospy.Publisher('excel_to_coordinate', String, queue_size=5)
  accout=getpass.getuser()
  excel_path= r'/home/%s/mapdata/cellmap.xlsx'%accout
  wb = load_workbook(filename = excel_path )
  ws=wb.get_sheet_by_name('parameters') 
  sheetnames=wb.get_sheet_names()
  sheet=wb.get_sheet_by_name(sheetnames[0])
  print "openning Worksheet titles:", ws.title,sheet.title
  print 'getting map resolution'
  resolution=ws.cell('C5').value
  print 'getting parameters'
  mapcentre_x=ws.cell('C11').value
  mapcentre_y=ws.cell('D11').value
  while not rospy.is_shutdown():
   pub.publish('start')
   try:
    excel=raw_input("waiting for input (字母+数字):\n")
    if excel=='':
     continue
   except:
    print 'try again'
    pass
   excel=list(excel)
   [col,row]=self.split(excel)
   print 'col:',col,' row:',row
   RequirePose=[self.column_to_number(col),row]
   print 'column number:',self.column_to_number(col)
   centremap=[int(round(mapcentre_x/resolution)),int(round(mapcentre_y/resolution))]
   print 'centremap[1]',centremap[1],'centremap[0]',centremap[0]
   height=sheet.get_highest_row()
   height+=centremap[1]
   print 'centremap',(2-centremap[0]),height
   mapcentre_column = '%s%s'%(get_column_letter(2-centremap[0]),height)

   rospy.loginfo('mapcentre position in excel position is ')
   rospy.loginfo(mapcentre_column)

   centre=[(2-centremap[0]),height]
   differ=[RequirePose[0]-centre[0],centre[1]-RequirePose[1]]
   column='%s,%s'%(differ[0]*resolution,differ[1]*resolution)
   print 'result: ',column
   for times in range(1):
    pub.publish(column)
   signal=raw_input('continue?:\n')
   if signal=='' or signal=='y' or signal=='yes':
    continue
   else:
    break

 def split(self,data):
  col=''
  row=''
  for i in data:
   if ord('Z')>=ord(i.upper())>=ord('A'):
    col+=i.upper()
   elif ord('9')>=ord(i.upper())>=ord('0'):
    row+=i.upper()
   else:
    pass
  return [col,int(row)]

 def column_to_number(self,col):
  rospy.loginfo('transfering...')
  number=0
  letters=list(col)
  size=len(col)
  for i in range(size):
   if letters[i] in string.ascii_letters:
    number=(ord(letters[i].upper())-64)+number*26**i
   else:
    number='NULL'
  return number

if __name__=='__main__':
 try:
  print "initialization system"
  excel_to_coordinate()
  print "process done and quit"
 except rospy.ROSInterruptException:
  rospy.loginfo("robot excel_to_coordinate node terminated.")
