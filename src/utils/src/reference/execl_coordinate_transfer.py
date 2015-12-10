#!/usr/bin/env python
#coding=utf-8
""" 

- Version 2.0 2015/9/11

this code transfer execl fram to coordinate fram

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

from openpyxl.cell import get_column_letter
import string
from openpyxl import load_workbook
import getpass,quat_to_angle,string

def excel2coordinate(excel):
 accout=getpass.getuser()
 excel_path= r'/home/%s/mapdata/cellmap.xlsx'%accout
 wb = load_workbook(filename = excel_path)
 ws=wb.get_sheet_by_name('parameters') 
 resolution=ws.cell('C5').value
 mapcentre_x=ws.cell('C11').value
 mapcentre_y=ws.cell('D11').value
 height=ws.cell('C8').value
 centremap=[1-int(round(mapcentre_x/resolution)),height+int(round(mapcentre_y/resolution))]
 [col,row]=split(list(excel))
 RequirePose=[column_to_number(col),int(row)]
 Require_Pose=[RequirePose[0]-centremap[0],RequirePose[1]-centremap[1]]
 column=[Require_Pose[0]*resolution,Require_Pose[1]*resolution]
 return column

def split(data):
 col=''
 row=''
 for i in data:
  if ord('Z')>=ord(i.upper())>=ord('A'):
   col+=i.upper()
  elif ord('9')>=ord(i.upper())>=ord('0'):
   row+=i.upper()
  else:
   pass
 return [col,int(row)]

def column_to_number(col):
 number=0
 letters=list(col)
 size=len(col)
 for i in range(size):
  if letters[i] in string.ascii_letters:
   number=(ord(letters[i].upper())-64)+number*26**i
  else:
   number='NULL'
 return number

def pointed_coordinate2execl(point):
 accout=getpass.getuser()
 excel_path= r'/home/%s/mapdata/cellmap.xlsx'%accout
 wb = load_workbook(filename = excel_path ) 
 ws=wb.get_sheet_by_name('parameters')
 mapcentre_x=ws.cell('C11').value
 mapcentre_y=ws.cell('D11').value
 resolution=ws.cell('C5').value
 height=ws.cell('C8').value
 centremap=[1-int(round(mapcentre_x/resolution)),height+int(round(mapcentre_y/resolution))]
 pointpose=[int(round(float(point[0])/resolution)),int(round(float(point[1])/resolution))]
 point_pose=[centremap[0]+pointpose[0],centremap[1]+pointpose[1]]
 pointed_column = [get_column_letter(point_pose[0]),point_pose[1]]
 return pointed_column

def coordinate2execl():
 accout=getpass.getuser()
 excel_path= r'/home/%s/mapdata/cellmap.xlsx'%accout
 wb = load_workbook(filename = excel_path ) 
 ws=wb.get_sheet_by_name('parameters')
 resolution=ws.cell('C5').value
 Pose_x=ws.cell('A16').value
 Pose_y=ws.cell('B16').value
 Pose_z=ws.cell('C16').value
 Rotation_x=ws.cell('A18').value
 Rotation_y=ws.cell('B18').value
 Rotation_z=ws.cell('C18').value
 Rotation_w=ws.cell('D18').value
 quaternion=[Rotation_x,Rotation_y,Rotation_z,Rotation_w]
 print 'robot pose:\n',[Pose_x,Pose_y,Pose_z],'\n',quaternion,'\n'
 mapcentre_x=ws.cell('C11').value
 mapcentre_y=ws.cell('D11').value
 height=ws.cell('C8').value
 centremap=[1-int(round(mapcentre_x/resolution)),height+int(round(mapcentre_y/resolution))]
 robotpose=[int(round(Pose_x/resolution)),int(round(Pose_y/resolution))]
 robot_pose=[centremap[0]+robotpose[0],centremap[1]+robotpose[1]]
 robot_column_pose = [get_column_letter(robot_pose[0]),robot_pose[1],quaternion]
 return robot_column_pose


