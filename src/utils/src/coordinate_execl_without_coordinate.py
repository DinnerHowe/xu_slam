#!/usr/bin/env python
#coding=utf-8

""" 

- Version 2.0 2015/10/16

before running this code pls install python-xlib python-xlrd python-xlwt python-openpyxl

this code transfer coordinate fram to execl fram
    

Copyright (c) 2015 Xu Zhihao (Howe).  All rights reserved.

This program is free software; you can redistribute it and/or modify

This programm is tested on kuboki base turtlebot. 

"""

import rospy
from openpyxl.cell import get_column_letter
from std_msgs.msg import String
from openpyxl import load_workbook
import getpass

class excel_to_coordinate():
 def __init__(self):
  rospy.init_node('excel_to_coordinate',anonymous=True)
  accout=getpass.getuser()
  excel_path= r'/home/%s/mapdata/cellmap.xlsx'%accout
  wb = load_workbook(filename = excel_path ) 
  print "Worksheet name(s):", wb.get_sheet_names()
  ws=wb.get_sheet_by_name('parameters')
  sheetnames=wb.get_sheet_names()
  sheet=wb.get_sheet_by_name(sheetnames[0])
  print "openning Worksheet title:", ws.title
  print 'getting map resolution'
  resolution=ws.cell('C5').value
  print 'resolution:',resolution
  print 'getting parameters','\n'
  Pose_x=ws.cell('A16').value
  Pose_y=ws.cell('B16').value
  Pose_z=ws.cell('C16').value
  Rotation_x=ws.cell('A18').value
  Rotation_y=ws.cell('B18').value
  Rotation_z=ws.cell('C18').value
  Rotation_w=ws.cell('D18').value
  Rotation=[Rotation_x,Rotation_y,Rotation_z,Rotation_w]
  mapcentre_x=ws.cell('C11').value
  mapcentre_y=ws.cell('D11').value

  while not rospy.is_shutdown():
   # map(0,0)
   centremap=[1-int(round(mapcentre_x/resolution)),1-int(round(mapcentre_y/resolution))]
   print 'distance between map (0,0) and A1:',centremap
   mapcentre_column = '%s%s'%(get_column_letter(centremap[0]),centremap[1])
   rospy.loginfo('mapcentre position in excel position is ')
   rospy.loginfo(mapcentre_column)

   #robot pose
   robotpose=[int(round(Pose_x/resolution)),int(round(Pose_y/resolution))]
   print 'distance between robotpose and map (0,0)',robotpose,'robotpose',[Pose_x,Pose_y],'orientation:',Rotation
   robot_pose=[centremap[0]+robotpose[0],centremap[1]+robotpose[1]]

   robot_column = '%s%s'%(get_column_letter(robot_pose[0]),robot_pose[1])
   rospy.loginfo('robot position in excel position is ')
   rospy.loginfo(robot_column)

   ###for testing show pointed position in data base
   try:
    col=raw_input("pls input col(x):\n")
    if col=='':
     continue
    row=raw_input('pls input row(y):\n')
    if row=='':
     continue
   except:
    pass
   print 'x:',col,' y:',row
   point=[col,row]

   pointpose=[int(round(float(point[0])/resolution)),int(round(float(point[1])/resolution))]
   print 'pointpose to centre point',pointpose

   point_pose=[centremap[0]+pointpose[0],centremap[1]+pointpose[1]]
   pointed_column = '%s%s'%(get_column_letter(point_pose[0]),point_pose[1])
   rospy.loginfo('pointed position in excel position is ')
   rospy.loginfo(pointed_column)

   robotpub = rospy.Publisher('robot_position_in_database', String, queue_size=5)
   robotpub.publish(robot_column)
   mapcentre = rospy.Publisher('mapcentre_in_database', String, queue_size=5)
   mapcentre.publish(mapcentre_column)
   pointed = rospy.Publisher('pointed_point_in_database', String, queue_size=5)
   pointed.publish(pointed_column)
   signal=raw_input('continue?:\n')
   if signal=='' or signal=='y' or signal=='yes':
    continue
   else:
    break

if __name__=='__main__':
 try:
  print "initialization system"
  excel_to_coordinate()
  print "process done and quit"
 except rospy.ROSInterruptException:
  rospy.loginfo("robot excel_to_coordinate node terminated.")
